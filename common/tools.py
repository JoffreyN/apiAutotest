import requests,os,json,time,datetime,logging,simplejson,traceback,sys,re
from common.database import queryCodelist,getDBconn,excuteSQL
from openpyxl import load_workbook
from jsonpath import jsonpath
from config import *
sys.path.append('..')

head={
	"Content-Type":"application/json;charset=UTF-8",
	'Connection':'close',
	'RHINO-PROXY-TO':'0.0.0.0:8080',
}

def getCellStr(cell,noStrip=0):
	if cell.value==None:return cell.value
	try:
		strs=str(cell.value) if noStrip else str(cell.value).strip()
	except AttributeError:
		strs=str(cell.value)
	return strs

def readExcel(fileName):
	if isinstance(fileName,list):
		try:
			fileName=[i for i in fileName if '.xlsx' in i][0]
		except IndexError:
			return [0,0,0]
	excelData=[]
	excel=load_workbook(f"testData/{fileName}")
	table=excel.active
	for row in table.iter_rows(min_row=2):
		lineDic={
			"id":getCellStr(row[0]),
			"desc":getCellStr(row[1]),
			"type":getCellStr(row[2]),
			"host":getCellStr(row[3]),
			"reqPath":getCellStr(row[4]),
			"reqType":getCellStr(row[5]),
			"reqHeader":getCellStr(row[6]),
			"dataType":getCellStr(row[7]),
			"reqData":getCellStr(row[8]),
			"checkType":getCellStr(row[9]),
			"checkMode":getCellStr(row[10]),
			"checkPoint":getCellStr(row[11],1),
			"correlation":getCellStr(row[12]),
			"skip":getCellStr(row[13]),
			"proxy":getCellStr(row[14]),
		}
		excelData.append(lineDic)
	excel.close()
	return excelData

def getDayMoneyFlow():
	head={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.102 Safari/537.36','Connection':'Close'}
	url='http://0.0.0.0/quotes_service/api/jsonp.php/var%20liveDateTableList=/HK_MoneyFlow.getDayMoneyFlowOtherInfo'
	logging.info(f'对比数据来源: {url}')
	resp=requests.get(url,headers=head)
	respTxt=resp.content.decode('unicode-escape')
	respJson=eval(re.findall(r'\{.+\}',respTxt)[0])
	flowInData={
		'HK2SH':respJson["south_hk_sh"]["daliyInflow"],
		'HK2SZ':respJson["south_hk_sz"]["daliyInflow"],
		'SH2HK':respJson["north_sh"]["daliyInflow"],
		'SZ2HK':respJson["north_sz"]["daliyInflow"],
	}
	# print(f'南下 港股通(沪) 当日资金流入: {respJson["south_hk_sh"]["daliyInflow"]}\n南下 港股通(深) 当日资金流入:  {respJson["south_hk_sz"]["daliyInflow"]}')
	# print(f'北上 沪股通 当日资金流入: {respJson["north_sh"]["daliyInflow"]}\n北上 深股通 当日资金流入:  {respJson["north_sz"]["daliyInflow"]}')
	return flowInData
	

def saveTodatabase(sql):
	conn=getDBconn(dataBaseInfo['test']['host'],dataBaseInfo['test']['uname'],dataBaseInfo['test']['pword'],'mysql')
	excuteSQL(conn,sql,0)

def formatSqlStr(strs):
	return str(strs).replace('"',"'").replace("'","''").replace("\\","")


def moveFiles(reportName,ops='move'):
	if platform=='darwin':
		cmd=f'mv {os.path.join(genPath,reportName)} {webPath}'
	elif platform=='win32':
		cmd=f'echo d | xcopy /r /y /e "{os.path.join(genPath,reportName)}" "{os.path.join(webPath,reportName)}"'
	elif platform=='linux':
		cmd=''
	os.system(cmd)

def postReq(path,dataJson,env,url=None,mod='post',nolog=0,qotToken=None,headerTraceLog=None):
	requests.packages.urllib3.disable_warnings()
	if not url:url=f'{domainJY[env]}{path}'
	if not nolog:logging.info(f'请求地址: {url}')
	if qotToken:head['cmbi-qot-token']=qotToken
	if headerTraceLog:head['headerTraceLog']=headerTraceLog
	# logging.info(f'head: {head}')
	if mod=='post':
		resp=requests.post(url,headers=head,json=dataJson,verify=False)
	elif mod=='get':
		resp=requests.get(url,headers=head,params=dataJson,verify=False)
	try:
		respJson=resp.json()
	except (simplejson.errors.JSONDecodeError,IndexError):
		respJson=resp.text
		logging.info(f'请求数据:{dataJson} 返回数据异常：{resp.text}')
		# raise Exception('解析返回json失败')
	return respJson

def getStockdata(stockCode,mkCode,source='gf'):
	if source=='gf':
		hostDic={
			161:'202.82.62.39:9080',
			101:'q3.gf.com.cn:9080',
			105:'q3.gf.com.cn:9080',
			302:'202.82.62.39:9080',
			303:'202.82.62.39:9080'
		}
		url=f'http://{hostDic[mkCode]}/v1/realtime'
		dataJson={
			"id":{"exchange":mkCode,"code":stockCode},
			"scene":0,
			"subscribe":False
		}
		logging.info(f'请求广发数据: {dataJson}')
		return postReq(0,dataJson,0,url=url)
	elif source=='jl':
		dataJson={'code':f'{mkCode}{stockCode}'}
		url=f'{domainJL["uat"]}/app/price/stockol'
		logging.info(f'请求捷力数据: {dataJson}')
		return postReq(0,dataJson,0,url=url,mod='get')
	elif source=='ft':
		market_code=f'{marketCodeNameDic[mkCode]}.{stockCode}'
		logging.info(f'请求富途数据: {market_code}')
		url=f'http://0.0.0.0:8083/snapshot/'
		dataJson={'market_code':market_code}
		return postReq(0,dataJson,0,url=url,mod='get')

def getStockList(mkCode):
	filePath=f"testData/{mkCode}_{time.strftime('%Y%m%d')}.txt"
	if os.path.exists(filePath):
		logging.info('从本地读取广发 stockData 数据')
		respJson=eval(open(filePath,'r',encoding='utf-8').read())
	else:
		logging.info('从广发获取 stockData 数据')
		url='http://0.0.0.0:9080/v1/codelist'
		dataJson={"Ver":1,"markets":[{"exchange":mkCode}]}
		respJson=postReq(0,dataJson,0,url=url)
		with open(filePath,'w',encoding='utf-8') as file:file.write(str(respJson))
	return respJson

def genCodeData(lens=3,datasource='database',env='test'):
	codeDataPath=f"testData/codeData_{time.strftime('%Y%m%d%H%M')}.txt"
	if os.path.exists(codeDataPath):
		logging.info('从本地读取 codeData 数据')
		codeData=eval(open(codeDataPath,'r',encoding='utf-8').read())
	else:
		logging.info('自动生成 codeData 数据')
		from testData.data import codeData
		stockDataDic={}
		HKlist,HKNlist,HKXlist,HKZlist,HKGlist,USlist,SHlist,SZlist=[],[],[],[],[],[],[],[]
		if datasource=='GF':
			for mkCode in [161,101,105,303,302,]:
				stockData=getStockList(mkCode)
				for i in stockData['markets'][0]['slots']:
					try:secus=i['secus']
					except KeyError:continue
					for j in secus:
						try:type2=j['type2']
						except KeyError:
							if mkCode in [302,303]:USlist.append({'types':'US','code':code,'name':name,'mkCode':mkCode})
							else:continue
						code=j['code'];name=j['secu_names'][0]['name']
						if type2==1003:# 普通
							HKlist.append({'types':'HK','code':code,'name':name,'mkCode':mkCode})
						elif type2==2001:# 牛证 
							HKNlist.append({'types':'HKN','code':code,'name':name,'mkCode':mkCode})
						elif type2==2002:# 熊证 
							HKXlist.append({'types':'HKX','code':code,'name':name,'mkCode':mkCode})
						elif type2==1402:# 认购 
							HKZlist.append({'types':'HKZ','code':code,'name':name,'mkCode':mkCode})
						elif type2==1403:# 认沽 
							HKGlist.append({'types':'HKG','code':code,'name':name,'mkCode':mkCode})
						elif type2==1001:
							if mkCode==101:# SH
								SHlist.append({'types':'SH','code':code,'name':name,'mkCode':mkCode})
							elif mkCode==105:# SZ
								SZlist.append({'types':'SZ','code':code,'name':name,'mkCode':mkCode})
			from random import sample
			for stockList in [HKlist,HKNlist,HKXlist,HKZlist,HKGlist,USlist,SHlist,SZlist]:
				codeData.extend(sample(stockList,lens))
		elif datasource=='database':
			logging.info('从数据库获取 stockData 数据')
			mk_types={161:'HK',101:'SH',105:'SZ',303:'US',302:'US',}
			# for mkCode in [161,101,105,303,302,]:
			for mkCode in [303,302]:
			# for mkCode in [161,101,105,303,302]:
				if mkCode==161:
					halfLens=int(lens/2)
					idLimitList=[(lens,34079,37491),(halfLens,37492,44187),(halfLens,44523,45360),(halfLens,45361,53921)]
					for i in idLimitList:
						codeData.extend([{'types':mk_types[mkCode],'code':i[0],'name':i[1],'mkCode':mkCode} for i in queryCodelist(marketCodeDic[mkCode],i[0],env=env,idLimit=i[1:])])
				else:
					codeData.extend([{'types':mk_types[mkCode],'code':i[0],'name':i[1],'mkCode':mkCode} for i in queryCodelist(marketCodeDic[mkCode],lens,env=env)])
		with open(codeDataPath,'w',encoding='utf-8') as file:file.write(str(codeData))
	# logging.info(f'codeData: {codeData}')
	# os.system('pause')
	return list(map(lambda x:list(x.values()),codeData))

def getToken(uname,pword,env,fromRedis=0):
	if fromRedis:
		logging.info('从Redis获取cookie信息...')
		import redis
		_dic_redis_port={'uat':6379,'test':6380,'prod':6381}
		r=redis.Redis(host='127.0.0.1',port=_dic_redis_port[env],decode_responses=True)
		session={
			'token':r.get('token'),
			'sessionId':r.get('sessionId'),
			'accountId':r.get('accountId'),
			'acctType':r.get('acctType'),
			'aecode':r.get('aecode'),
			'marginMax':r.get('marginMax'),
		}
		if checkSession(session,env):return session
		else:logging.info('Redis内cookie信息已失效，重新登录...')
	else:
		try:
			sessionDic=eval(open('testData/sessions','r',encoding='utf-8').read())
			try:
				session=sessionDic[env][uname]
				logging.info('从本地获取cookie信息...')
				if checkSession(session,env):return session
				else:logging.info('本地获取cookie信息已失效，重新登录...')
			except KeyError:
				logging.info(f'本地无 {uname} cookie信息，自动登录获取...')
		except FileNotFoundError:
			sessionDic={'test':{},'uat':{},'prod':{}}

	url=f'{domainCMBI[env]}/app/user/login'
	key={'account':uname,'password':pword}
	head={
		'Connection':'close',
		'Content-Type':'application/x-www-form-urlencoded',
		'User-Agent':'zyapp/2.2.1.36591 (HONOR COLAL10; Android 9) uuid/VBJDU19510007442 channel/Atest1 redgreensetting/red language/zhCN versionCode/33562625'
	}
	# print('登录地址:',url)
	for i in range(5):
		resp=requests.post(url,headers=head,data=key)
		respJson=resp.json()
		# print('登录返回结果:',respJson)
		# jsonpath_getOne(dataJson,key,dataType='float',nan=0,path=None)
		session={
			"token":jsonpath_getOne(respJson,'token',dataType='str',nan='0'),
			"sessionId":jsonpath_getOne(respJson,'sessionid','str'),
			"accountId":jsonpath_getOne(respJson,'accountid','str'),
			"acctType":jsonpath_getOne(respJson,'acctype','str'),
			"aecode":jsonpath_getOne(respJson,'aecode','str'),
			"marginMax":jsonpath_getOne(respJson,'margin_max','str'),
		}
		if session['token']=='0':
			logging.info(f'登录失败: {respJson} 3秒后重试...')
			time.sleep(3)
		else:
			break

	if fromRedis:
		r.set('token',session['token'])
		r.set('sessionId',session['sessionId'])
		r.set('acctType',session['acctType'])
		r.set('aecode',session['aecode'])
		r.set('marginMax',session['marginMax'])
		r.set('accountId',session['accountId'])
		r.set('accountName',jsonpath_getOne(respJson,'account_name','str'))
		r.set('operatorNo',jsonpath_getOne(respJson,'user_id','str'))
		r.close()
	else:
		sessionDic[env][uname]=session
		with open('testData/sessions','w',encoding='utf-8') as file:file.write(str(sessionDic))
	return session

def checkSession(session,env):
	url=f'{domainCMBI[env]}/gateway/order/stockMarketRole'
	# resp=requests.post(url,headers=head,data=key)
	respJson=postReq(0,session,env,url=url,nolog=1)
	return respJson['success']

# def saveCookie(cookie,file):
# 	sys.path.append('..')
# 	with open(f'testData/cookie/{file}','w',encoding='utf-8') as f:
# 		f.write(cookie)

def assertCom_ws(code,dataJson,logit=0):
	if code==1001:# InitConnect
		pass
	elif code==3012: #Qot_GetOrderBook
		assertObj.assertGreaterEqual(jsonpath_getOne(dataJson,0,path='$..buy..price',nan=0)+jsonpath_getOne(dataJson,0,path='$..sell..price',nan=0),0)
		assertObj.assertGreaterEqual(jsonpath_getOne(dataJson,0,path='$..buy..volume',nan=0)+jsonpath_getOne(dataJson,0,path='$..sell..volume',nan=0),0)
	elif code==3005: #Qot_GetSnapshotQot
		market=jsonpath_getOne(dataJson,'market','str')
		stockcode=jsonpath_getOne(dataJson,'code','str')
		# assertObj.assertIsNotNone(jsonpath_getOne(dataJson,'curPrice',nan=None))
		# assertObj.assertIsNotNone(jsonpath_getOne(dataJson,'highPrice',nan=None))
		# assertObj.assertIsNotNone(jsonpath_getOne(dataJson,'lowPrice',nan=None))
		# assertObj.assertIsNotNone(jsonpath_getOne(dataJson,'openPrice',nan=None))
		assertObj.assertIsNotNone(jsonpath_getOne(dataJson,'lastClosePrice',nan=None))
		if logit:logging.info(f'snapshot {market}{stockcode} 测试通过')
	elif code==3008:# Qot_GetTimeShare
		market=jsonpath_getOne(dataJson,'market','str')
		stockcode=jsonpath_getOne(dataJson,'code','str')
		minutes=len(jsonpath_getOne(dataJson,'rtList',None,nan=''))
		logging.info(f"{jsonpath_getOne(dataJson,'code','str')} 分时个数: {minutes}")
		assertObj.assertGreater(minutes,0)
		if logit:logging.info(f'TimeShare {market}{stockcode} 测试通过')
	elif code==3001:# Qot_Sub
		assertObj.assertEqual('subscribe success',dataJson['retMsg'])
	elif code==3301:# Qot_UpdateSecuritySnapshot
		# assertObj.assertEqual('ok',dataJson['retMsg'])
		assertObj.assertIsNotNone(jsonpath_getOne(dataJson,'updateTime','str'))
		assertObj.assertIsNotNone(jsonpath_getOne(dataJson,'security','str'))
	elif code==3302:# Qot_UpdateOrderBook
		pass
	elif code==3303:# Qot_UpdateBasicQot
		pass

def jsonpath_getOne(dataJson,key,dataType='float',nan=0,path=None):
	if not path:path=f'$..{key}'
	try:value=jsonpath(dataJson,path)[0]
	except TypeError:value=nan
	if dataType=='float':
		try:value=float(value)
		except TypeError:pass
	return value

def jsonpath_getAll(dataJson,key=None,path=None):
	if not key and not path:return 'key 或 paht 至少要传1个'
	if not path:path=f'$..{key}'
	results=jsonpath(dataJson,path)
	if not results:results=[]
	return results

def sendMail(text,platformName,fromInfo=None,cusReceiver=None):
	import smtplib
	from email.header import Header
	from email.mime.text import MIMEText
	from email.utils import formataddr
	if cusReceiver:
		receiver,mailToCc=cusReceiver.split(),[]
	else:
		from config import receiver,mailToCc
	sender='****@****.***.**'
	subject=f'【{platformName}】壹隆环球接口自动化测试报告'

	msg=MIMEText(text,'html','utf-8')
	msg['Subject']=Header(subject,'utf-8') #设置主题和格式
	if not fromInfo:fromInfo='行情接口监控 --【失败告警】'
	msg['From']=formataddr([fromInfo,sender])
	msg['To']=';'.join(receiver)
	msg['Cc']=';'.join(mailToCc)
	
	smtp=smtplib.SMTP('****.****.***.**',25)
	# smtp=smtplib.SMTP_SSL('smtp.qq.com',465)
	# smtp.ehlo()
	# smtp.starttls()
	# smtp.login(username, password)
	smtp.sendmail(sender, receiver+mailToCc, msg.as_string())
	smtp.quit()

def isTradeTime(timeNow=None,market='HK'):
	#传入时间戳
	from chinese_calendar import is_workday
	if not timeNow:timeNow=time.time()
	date,weekDay=time.strftime('%Y%m%d %w').split()
	resp=is_workday(datetime.date(int(date[:4]),int(date[4:6]),int(date[6:])))# 节假日返回 False
	if (not resp) or (weekDay in ['0','6']):#周六日休市
		return 0
	else:
		if market=='HK':
			start1=time.mktime(time.strptime(f'{date} 09:30:00','%Y%m%d %X'))
			stop1=time.mktime(time.strptime(f'{date} 12:00:00','%Y%m%d %X'))
			start2=time.mktime(time.strptime(f'{date} 13:00:00','%Y%m%d %X'))
			stop2=time.mktime(time.strptime(f'{date} 16:00:00','%Y%m%d %X'))
		elif market=='A':
			start1=time.mktime(time.strptime(f'{date} 09:30:00','%Y%m%d %X'))
			stop1=time.mktime(time.strptime(f'{date} 11:30:00','%Y%m%d %X'))
			start2=time.mktime(time.strptime(f'{date} 13:00:00','%Y%m%d %X'))
			stop2=time.mktime(time.strptime(f'{date} 15:00:00','%Y%m%d %X'))
		elif market=='US':
			start1=time.mktime(time.strptime(f'{date} 21:30:00','%Y%m%d %X'))
			stop1=start1+6.5*60*60
			start2=start1
			stop2=stop1
		timeNow_strf=time.strftime('%Y-%m-%d %X',time.localtime(timeNow))
		if start1<timeNow<stop1 or start2<timeNow<stop2:
			logging.info(f'当前时间 {timeNow_strf} 为 {market} 交易时间')
			return 1
		else:
			logging.info(f'当前时间 {timeNow_strf} 为 {market} 菲交易时间')
			return 0
